from django.shortcuts import render, redirect
from .models import Employee, Order, Serial, Path, MoveHistory, DeleteHistory
from django.http import JsonResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill

def index(request):
    context = {
    }
    return render(request, 'index.html', context)

def track_serial(request):
    alert = 'NOTFOUND'
    order_no = request.GET.get('order_no')
    serial_code = request.GET.get('serial_code')
    print(order_no, serial_code)
    drawing_no = '-'
    fg_code = '-'
    creator = '-'
    path_length = 0
    path_status = []
    path_emp_id = []
    path_location = []
    path_time_stamp = []
    order_is_exist = Order.objects.filter(no=order_no).exists()
    if(order_is_exist == True):
        order = Order.objects.get(no=order_no)
        order_no = order.no
        drawing_no = order.drawing_no
        fg_code = order.fg_code
        creator = order.emp_id
        serial_is_exist = Serial.objects.filter(code=serial_code,order=order).exists()
        if(serial_is_exist == True):
            alert = 'FOUND'
            serial = Serial.objects.get(code=serial_code,order=order)
            #-- GET ALL PATH
            paths = Path.objects.filter(serial=serial).order_by('date_published')
            path_length = len(paths)
            for path in paths:
                path_status.append(path.status)
                path_emp_id.append(path.emp_id)
                path_location.append(path.location)
                path_time_stamp.append(path.date_published.strftime("%b. %d, %Y, %I:%M %p"))
    data = {
        'alert': alert,
        'serial_code' : serial_code,
        'order_no' : order_no,
        'drawing_no' : drawing_no,
        'fg_code' : fg_code,
        'creator' : creator,
        'path_length' : path_length,
        'path_status' : path_status,
        'path_emp_id' : path_emp_id,
        'path_location' : path_location,
        'path_time_stamp' : path_time_stamp,
    }
    return JsonResponse(data)

def change_location(request):
    context = {
    }
    return render(request, 'change_location.html', context)

def exc_change_location(request):
    alert = 'NOTFOUND'
    order_no = request.GET.get('order_no')
    serial_code = request.GET.get('serial_code')
    emp_id = request.GET.get('emp_id')
    location = request.GET.get('location')
    order_is_exist = Order.objects.filter(no=order_no).exists()
    if(order_is_exist == True):
        order = Order.objects.get(no=order_no)
        serial_is_exist = Serial.objects.filter(code=serial_code,order=order).exists()
        if(serial_is_exist == True):
            alert = 'REJECTED'
            serial = Serial.objects.get(code=serial_code)
            last_path = Path.objects.filter(serial=serial).order_by('-date_published')[0]
            if(last_path.status != 'REJECTED'):
                #-- SAVE PATH
                alert = 'SUCCESS'
                status = 'RECEIVED'
                path_new = Path(serial=serial,emp_id=emp_id,location=location,status=status)
                path_new.save()
    data = {
        'alert': alert,
    }
    return JsonResponse(data)

def new_workorder(request):
    context = {
    }
    return render(request, 'new_workorder.html', context)

def validate_order_no(request):
    order_no = request.GET.get('order_no')
    data = {
        'is_exist': Order.objects.filter(no=order_no).exists()
    }
    return JsonResponse(data)

def validate_serial_code(request):
    order_no = request.GET.get('order_no')
    serial_code = request.GET.get('serial_code')
    order = Order.objects.get(no=order_no)
    data = {
        'is_exist': Serial.objects.filter(code=serial_code,order=order).exists()
    }
    return JsonResponse(data)

def new_workorder_save(request):
    # COLLECT DATA
    order_no = request.POST.get('order_no')
    drawing_no = request.POST.get('drawing_no')
    fg_code = request.POST.get('fg_code')
    emp_id = request.POST.get('emp_id')
    location = request.POST.get('location')
    serial_code_list = []
    serial_code_count = int(request.POST.get('serial_code_count'))
    for i in range(serial_code_count):
        serial_code = request.POST.get('serial_code_' + str(i))
        serial_code_list.append(serial_code)
    # PREPARE DATA
    status = 'CREATED'
    # SAVE DATA
    #-- WORK ORDER
    order_new = Order(no=order_no,drawing_no=drawing_no,fg_code=fg_code,emp_id=emp_id)
    order_new.save()
    for i in range(serial_code_count):
        #-- SERIAL CODE
        serial_new = Serial(code=serial_code_list[i],order=order_new)
        serial_new.save()
        #-- PATH
        path_new = Path(serial=serial_new,emp_id=emp_id,location=location,status=status)
        path_new.save()
    context = {
    }
    return redirect('/order/' + order_new.no)

def order_master(request):
    orders = Order.objects.all()
    items = []
    for order in orders:
        serials = Serial.objects.filter(order=order)
        items.append(len(serials))
    context = {
        'orders' : orders,
        'items' : items,
    }
    return render(request, 'order_master.html', context)

def emp_master(request):
    context = {
    }
    return render(request, 'emp_master.html', context)

def fg_master(request):
    context = {
    }
    return render(request, 'fg_master.html', context)

def order(request, order_no):
    orders = Order.objects.all().order_by('-date_published')
    order = Order.objects.get(no=order_no)
    serial_codes = Serial.objects.filter(order=order)
    path_list = []
    last_status = []
    last_receivers = []
    last_locations = []
    last_timestamps = []
    for serial_code in serial_codes:
        paths = Path.objects.filter(serial=serial_code).order_by('date_published')
        path_list.append(paths)
        if len(paths) > 0:
            last_path = paths[len(paths) - 1]
            last_status.append(last_path.status)
            last_receivers.append(last_path.emp_id)
            last_locations.append(last_path.location)
            last_timestamps.append(last_path.date_published)
        else:
            last_status.append('-')
            last_receivers.append('-')
            last_locations.append('-')
            last_timestamps.append('-')
    context = {
        'order' : order,
        'serial_codes' : serial_codes,
        'path_list' : path_list,
        'last_status' : last_status,
        'last_receivers' : last_receivers,
        'last_locations' : last_locations,
        'last_timestamps' : last_timestamps,
        'orders' : orders,
    }
    return render(request, 'order.html', context)

def create_serial(request):
    # COLLECT DATA
    order_no = request.GET.get('order_no')
    serial_code = request.GET.get('serial_code')
    emp_id = request.GET.get('emp_id')
    location = request.GET.get('location')
    # PREPARE DATA
    order = Order.objects.get(no=order_no)
    status = 'CREATED'
    # SAVE DATA
    #-- SERIAL CODE
    serial_new = Serial(code=serial_code,order=order)
    serial_new.save()
    #-- PATH
    path_new = Path(serial=serial_new,emp_id=emp_id,location=location,status=status)
    path_new.save()
    data = {
    }
    return JsonResponse(data)

def multiple_receive(request):
    # COLLECT DATA
    serial_code_list = request.GET.getlist('serial_code_list[]')
    emp_id = request.GET.get('emp_id')
    location = request.GET.get('location')
    # PREPARE DATA
    status = 'RECEIVED'
    # SAVE DATA
    for i in range(len(serial_code_list)):
        serial = Serial.objects.get(id=serial_code_list[i])
        path_new = Path(serial=serial,emp_id=emp_id,location=location,status=status)
        path_new.save()
    data = {
    }
    return JsonResponse(data)

def add_po_no(request):
    # COLLECT DATA
    serial_code_list = request.GET.getlist('serial_code_list[]')
    po_no = request.GET.get('po_no')
    # SAVE DATA
    for i in range(len(serial_code_list)):
        serial = Serial.objects.get(id=serial_code_list[i])
        serial.po_no = po_no
        serial.save()
    data = {
    }
    return JsonResponse(data)

def serial_receive(request):
    # COLLECT DATA
    serial_code_id = request.GET.get('serial_code_id')
    emp_id = request.GET.get('emp_id')
    location = request.GET.get('location')
    # PREPARE DATA
    serial = Serial.objects.get(id=serial_code_id)
    status = 'RECEIVED'
    # SAVE DATA
    path_new = Path(serial=serial,emp_id=emp_id,location=location,status=status)
    path_new.save()
    data = {
    }
    return JsonResponse(data)

def serial_reject(request):
    # COLLECT DATA
    serial_code_id = request.GET.get('serial_code_id')
    emp_id = request.GET.get('emp_id')
    location = request.GET.get('location')
    # PREPARE DATA
    serial = Serial.objects.get(id=serial_code_id)
    status = 'REJECTED'
    # SAVE DATA
    path_new = Path(serial=serial,emp_id=emp_id,location=location,status=status)
    path_new.save()
    data = {
    }
    return JsonResponse(data)

def serial_edit(request):
    # COLLECT DATA
    serial_code_id = request.GET.get('serial_code_id')
    new_serial_code = request.GET.get('new_serial_code')
    # PREPARE DATA
    serial = Serial.objects.get(id=serial_code_id)
    # SAVE DATA
    serial.code = new_serial_code
    serial.save()
    data = {
    }
    return JsonResponse(data)

def serial_move(request):
    # COLLECT DATA
    serial_code_id = request.GET.get('serial_code_id')
    order_no = request.GET.get('order_no')
    emp_id = request.GET.get('emp_id')
    reason = request.GET.get('reason')
    # PREPARE DATA
    serial = Serial.objects.get(id=serial_code_id)
    order = Order.objects.get(no=order_no)
    move_h_new = MoveHistory(serial_code=serial.code,old_order_no=serial.order.no,new_order_no=order_no,emp_id=emp_id,reason=reason)
    move_h_new.save()
    # SAVE DATA
    serial.order = order
    serial.save()
    data = {
    }
    return JsonResponse(data)

def serial_delete(request):
    # COLLECT DATA
    serial_code_id = request.GET.get('serial_code_id')
    emp_id = request.GET.get('emp_id')
    reason = request.GET.get('reason')
    # PREPARE DATA
    serial = Serial.objects.get(id=serial_code_id)
    paths = Path.objects.filter(serial=serial).order_by('date_published')
    path_history = ''
    for path in paths:
        path_history = path_history + '- ' + str(path.date_published) + ' @ ' + path.location + ' BY ' + path.emp_id + ' (' + path.status + ')\n'
    del_h_new = DeleteHistory(serial_code=serial.code,order_no=serial.order.no,emp_id=emp_id,reason=reason,path_history=path_history)
    del_h_new.save()
    # SAVE DATA
    serial.delete()
    data = {
    }
    return JsonResponse(data)

def import_data(request):
    wb = load_workbook(filename = 'media/data.xlsx')
    ws = wb.worksheets[0]
    skip_count = 2
    for i in range(ws.max_row + 1):
        if i < skip_count:
            continue
        fg_code = ws['A' + str(i)].value
        order_no = ws['B' + str(i)].value
        serial_code = ws['C' + str(i)].value
        drawing_no='-'
        emp_id = '5731'
        location = '-'
        status = 'CREATED'
        #-- WORK ORDER
        order_is_exist = Order.objects.filter(no=order_no).exists()
        if(order_is_exist == False):
            order_new = Order(no=order_no,drawing_no=drawing_no,fg_code=fg_code,emp_id=emp_id)
            order_new.save()
        order = Order.objects.get(no=order_no)
        serial_is_exist = Serial.objects.filter(code=serial_code,order=order).exists()
        if(serial_is_exist == False):
            #-- SERIAL CODE
            serial_new = Serial(code=serial_code,order=order)
            serial_new.save()
            #-- PATH
            path_new = Path(serial=serial_new,emp_id=emp_id,location=location,status=status)
            path_new.save()
    return redirect('/new_workorder/')
